import glob, os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import locale


class Update_Excel():
    """
    Updates the excel files for the provinces in
    a specified region with the specified data.

    Parameters
    ----------
    path : string, default=None
        The location of the folder containing
        the excel files.

    regione : string, {""}
        The name of the region to update
    """

    def __init__(self, region, path=None):

        self.path = path
        self.region = region
        self.region_list = ["Abruzzo", "Basilicata", "Calabria", "Emilia-Romagna", "Liguria", "Molise", "Sicilia",
                            "Toscana", "Umbria"]
        locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')

    def _get_default_path(self):
        """
        Get the corresponding default path for the
        excel files of the specified region.
        """
        if self.region == "Abruzzo":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Abruzzo"
        elif self.region == "Basilicata":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Basilicata"
        elif self.region == "Calabria":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Calabria"
        elif self.region == "Emilia-Romagna":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Emilia-Romagna"
        elif self.region == "Liguria":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Liguria"
        elif self.region == "Molise":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Molise"
        elif self.region == "Sicilia":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Sicilia"
        elif self.region == "Toscana":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Toscana"
        elif self.region == "Umbria":
            return "C:/Users/Rijin/Documents/STAGEshidash_COVID_Italy/morti_data/Morti_Umbria"
        else:
            raise ValueError(self.region + " is not valid. " \
                                           "Please choose among " + ' '.join([str(elem) for elem in self.region_list]))

    def _get_lista_prov(self):
        """
        Get the corresponding list of provinces
        from the specified region.
        """
        if self.region == "Abruzzo":
            return ["Chieti", "L'Aquila", "Pescara", "Teramo"]
        elif self.region == "Basilicata":
            return ["Matera", "Potenza"]
        elif self.region == "Calabria":
            return ["Catanzaro", "Cosenza", "Crotone", "Reggio di Calabria", "Vibo Valentia"]
        elif self.region == "Emilia-Romagna":
            return ["Bologna", "Ferrara", "Forlì-Cesena", "Modena", "Parma", "Piacenza", "Ravenna",
                    "Reggio nell'Emilia", "Rimini"]
        elif self.region == "Liguria":
            return ["Imperia","Savona"]
        elif self.region == "Molise":
            return ["Campobasso", "Isernia"]
        elif self.region == "Sicilia":
            return ["Agrigento", "Caltanissetta", "Catania", "Enna", "Messina", "Palermo", "Ragusa", "Siracusa",
                    "Trapani"]
        elif self.region == "Toscana":
            return ["Arezzo", "Firenze", "Grosseto", "Livorno", "Lucca", "Massa Carrara", "Pisa", "Pistoia", "Prato",
                    "Siena"]
        elif self.region == "Umbria":
            return ["Perugia", "Terni"]
        else:
            raise ValueError(self.region + " is not valid. " \
                                           "Please choose among " + ' '.join([str(elem) for elem in self.region_list]))

    def _get_files_list(self):
        """
        Enumerates the files in the folder and retrieves
        their full paths.

        Returns
        -------
        files : list
            The list of full paths for the files
            in the folder.
        """

        if self.path is None:
            self.path = self._get_default_path()

        if not str(self.path).endswith("/"):
            self.path = str(self.path) + "/"

        files = glob.glob(self.path + '*.xlsx')
        files.sort()

        return files

    def _get_decessi_casi(self, df, i, prov, csv_prov_github):
        """
        Get the number of deaths and infected

        Parameters
        ----------
        df : Pandas DataFrame
            The DataFrame containing the new data.

        i : int
            Counter

        prov : string
            The name of the province

        csv_prov_github : Pandas DataFrame
            The DataFrame containing data from the Protezione
            Civile.

        Returns
        -------
        tot_casi : int
            The cumulative number of infected.

        decessi : int
            The number of deaths. It can either be
            the daily or cumulative count.
        """

        tot_casi, decessi = None, None

        if self.region in ["Calabria", "Sicilia"]:
            # Casi_att_positivi
            tot_casi = csv_prov_github[csv_prov_github["denominazione_provincia"] == prov]["totale_casi"].values[0]
            # decessi_tot
            decessi_tot = df.iloc[i]["decessi_tot"]
            return tot_casi, decessi_tot

        else:
            tot_casi = csv_prov_github[csv_prov_github["denominazione_provincia"] == prov]["totale_casi"].values[0]
            decessi = df.iloc[i]["decessi"]
            return tot_casi, decessi

    def update_xls(self, df, day_spec):
        """
        Updates the excel files with the specified data for all the provinces
        of the region

        Parameters
        ----------
        df : Pandas DataFrame
            The DataFrame containing the new data.
        """

        self.region = str(self.region)

        if self.region not in self.region_list:
            raise ValueError(self.region + " is not valid. " \
                                            "Please choose among " + ' '.join([str(elem) for elem in self.region_list]))

        self.prov_list = self._get_lista_prov()

        files = self._get_files_list()

        df = df.sort_values(by=["provincia"])

        date = datetime.strftime(datetime.now() - timedelta(day_spec), "%Y%m%d")

        # Scarica il csv province della PCM
        csv_prov_github = pd.read_csv(
            "https://raw.githubusercontent.com/pcm-dpc/COVID-19/master/dati-province/dpc-covid19-ita-province.csv",
            error_bad_lines=False)
        csv_prov_github['data'] = pd.to_datetime(csv_prov_github.data)
        csv_prov_github['data'] = csv_prov_github['data'].dt.strftime('%Y%m%d')
        csv_prov_github = csv_prov_github[(csv_prov_github['data'] == date)]

        if self.region in ["Calabria", "Sicilia"]:
            for i, file in enumerate(files):
                prov = self.prov_list[i]
                tot_casi, decessi_tot = self._get_decessi_casi(df, i, prov, csv_prov_github)
                self._write_xls_CnS(file, decessi_tot, tot_casi)
                print(prov + ": done")

        else:
            for i, file in enumerate(files):
                prov = self.prov_list[i]
                tot_casi, decessi = self._get_decessi_casi(df, i, prov, csv_prov_github)
                self._write_xls(file, decessi, tot_casi)
                print(prov + ": done")

    def _write_xls(self, file, decessi, tot_casi):
        """
        Updates the excel file for the specified province

        Parameters
        ----------
        file : string
            The file path.

        decessi : int
            The number of deaths. It can either be
            the daily or cumulative count.

        tot_casi : int
            The cumulative number of infected.
        """

        # carica il xlsx
        wb = load_workbook(file)

        # seleziona il foglio1
        sheet = wb.active

        # n di righe
        max_row = sheet.max_row

        # Province
        sheet.cell(row=max_row + 1, column=1).value = sheet.cell(row=max_row, column=1).value

        # Day
        if sheet.cell(row=max_row, column=2).value is not None:
            sheet.cell(row=max_row + 1, column=2).value = sheet.cell(row=max_row, column=2).value + 1
        else:
            sheet.cell(row=max_row + 1, column=2).value = max_row - 1

        # Data
        sheet.cell(row=max_row + 1, column=3).value = sheet.cell(row=max_row, column=3).value + timedelta(1)
        sheet.cell(row=max_row + 1, column=3).number_format = 'YYYY-MM-DD'

        # Casi_att_positivi
        sheet.cell(row=max_row + 1, column=5).value = tot_casi

        # Nuovi_casi
        sheet.cell(row=max_row + 1, column=4).value = "=E" + str(max_row + 1) + "-E" + str(max_row)

        # decessi
        sheet.cell(row=max_row + 1, column=6).value = decessi

        # decessi_tot
        sheet.cell(row=max_row + 1, column=7).value = "=G" + str(max_row) + "+F" + str(max_row + 1)

        # tasso_letal
        sheet.cell(row=max_row + 1, column=8).value = "=G" + str(max_row + 1) + "/E" + str(max_row + 1)
        sheet.cell(row=max_row + 1, column=8).number_format = '0.00%'

        # tasso_incidenza_100000
        form = str(sheet.cell(row=max_row, column=9).value)
        form = form[form.find("/") + 1:form.find("*")]
        sheet.cell(row=max_row + 1, column=9).value = "=E" + str(max_row + 1) + "/" + form + "*100000"

        # salva
        wb.save(file)

    def _write_xls_CnS(self, file, decessi_tot, tot_casi):

        # carica il xlsx
        wb = load_workbook(file)

        # seleziona il foglio1
        sheet = wb.active

        # n di righe
        max_row = sheet.max_row

        # Province
        sheet.cell(row=max_row + 1, column=1).value = sheet.cell(row=max_row, column=1).value

        # Day
        if sheet.cell(row=max_row, column=2).value is not None:
            sheet.cell(row=max_row + 1, column=2).value = sheet.cell(row=max_row, column=2).value + 1
        else:
            sheet.cell(row=max_row + 1, column=2).value = max_row - 1

        # Data
        sheet.cell(row=max_row + 1, column=3).value = sheet.cell(row=max_row, column=3).value + timedelta(1)
        sheet.cell(row=max_row + 1, column=3).number_format = 'YYYY-MM-DD'

        # Casi_att_positivi
        sheet.cell(row=max_row + 1, column=5).value = tot_casi

        # Nuovi_casi
        sheet.cell(row=max_row + 1, column=4).value = "=E" + str(max_row + 1) + "-E" + str(max_row)

        # decessi
        sheet.cell(row=max_row + 1, column=6).value = "=G" + str(max_row + 1) + "-G" + str(max_row)

        # decessi_tot
        sheet.cell(row=max_row + 1, column=7).value = decessi_tot

        # tasso_letal
        sheet.cell(row=max_row + 1, column=8).value = "=G" + str(max_row + 1) + "/E" + str(max_row + 1)
        sheet.cell(row=max_row + 1, column=8).number_format = '0.00%'

        # tasso_incidenza_100000
        form = str(sheet.cell(row=max_row, column=9).value)
        form = form[form.find("/") + 1:form.find("*")]
        sheet.cell(row=max_row + 1, column=9).value = "=E" + str(max_row + 1) + "/" + form + "*100000"

        # salva
        wb.save(file)